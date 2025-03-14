"""
Module xử lý trích xuất thông tin từ file Excel
"""

import os
import logging
import pandas as pd
import openpyxl
from typing import Dict, Any, List, Optional, Tuple, Union
from pathlib import Path

from app.core.logging import get_logger

logger = get_logger("cv_parser")


class ExcelParser:
    """
    Lớp trích xuất dữ liệu từ file Excel
    """

    def __init__(self):
        """
        Khởi tạo parser
        """
        pass

    def extract_text(self, file_path: str) -> str:
        """
        Trích xuất văn bản từ file Excel

        Args:
            file_path: Đường dẫn đến file Excel

        Returns:
            str: Văn bản đã trích xuất
        """
        try:
            # Đọc tất cả các sheet trong file Excel
            excel_data = pd.ExcelFile(file_path)
            all_text = []

            for sheet_name in excel_data.sheet_names:
                df = pd.read_excel(excel_data, sheet_name=sheet_name)

                # Lấy tên sheet
                all_text.append(f"--- Sheet: {sheet_name} ---")

                # Xử lý các dòng trống hoặc NaN
                df = df.fillna("")

                # Chuyển đổi tất cả cột thành chuỗi
                for col in df.columns:
                    df[col] = df[col].astype(str)

                # Thêm tên cột
                all_text.append(" | ".join([str(col) for col in df.columns]))

                # Thêm dữ liệu
                for _, row in df.iterrows():
                    row_text = " | ".join([str(val) for val in row.values])
                    all_text.append(row_text)

                all_text.append("\n")  # Thêm dòng trống giữa các sheet

            return "\n".join(all_text)
        except Exception as e:
            logger.error(f"Lỗi khi trích xuất văn bản từ Excel {file_path}: {str(e)}")
            return ""

    def extract_as_dataframes(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """
        Trích xuất dữ liệu dưới dạng DataFrame cho từng sheet

        Args:
            file_path: Đường dẫn đến file Excel

        Returns:
            Dict[str, pd.DataFrame]: Từ điển các DataFrame với key là tên sheet
        """
        try:
            excel_data = pd.ExcelFile(file_path)
            dataframes = {}

            for sheet_name in excel_data.sheet_names:
                df = pd.read_excel(excel_data, sheet_name=sheet_name)
                dataframes[sheet_name] = df

            return dataframes
        except Exception as e:
            logger.error(f"Lỗi khi trích xuất dataframes từ Excel: {str(e)}")
            return {}

    def extract_metadata(self, file_path: str) -> Dict[str, Any]:
        """
        Trích xuất metadata từ file Excel

        Args:
            file_path: Đường dẫn đến file Excel

        Returns:
            Dict[str, Any]: Metadata từ file Excel
        """
        metadata = {}
        try:
            # Sử dụng openpyxl để lấy metadata chi tiết hơn
            workbook = openpyxl.load_workbook(file_path, read_only=True)

            # Thông tin cơ bản
            metadata['sheet_names'] = workbook.sheetnames
            metadata['sheet_count'] = len(workbook.sheetnames)

            # Thuộc tính cốt lõi nếu có
            if workbook.properties:
                for attr in ['creator', 'lastModifiedBy', 'created', 'modified', 'title', 'subject', 'keywords',
                             'category']:
                    if hasattr(workbook.properties, attr):
                        metadata[attr] = getattr(workbook.properties, attr)

            # Thêm số hàng và cột mỗi sheet
            sheet_details = {}
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_details[sheet_name] = {
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column
                }

            metadata['sheet_details'] = sheet_details

            # Đóng workbook
            workbook.close()

        except Exception as e:
            logger.error(f"Lỗi khi trích xuất metadata: {str(e)}")

        return metadata

    def extract_structured_data(self, file_path: str, header_row: int = 0) -> Dict[str, List[Dict[str, Any]]]:
        """
        Trích xuất dữ liệu có cấu trúc từ file Excel

        Args:
            file_path: Đường dẫn đến file Excel
            header_row: Chỉ số hàng tiêu đề (mặc định là 0 - hàng đầu tiên)

        Returns:
            Dict[str, List[Dict[str, Any]]]: Từ điển dữ liệu có cấu trúc với key là tên sheet
        """
        structured_data = {}
        try:
            excel_data = pd.ExcelFile(file_path)

            for sheet_name in excel_data.sheet_names:
                df = pd.read_excel(excel_data, sheet_name=sheet_name, header=header_row)

                # Xử lý các giá trị NaN
                df = df.where(pd.notna(df), None)

                # Chuyển DataFrame thành danh sách từ điển
                records = df.to_dict('records')
                structured_data[sheet_name] = records

            return structured_data
        except Exception as e:
            logger.error(f"Lỗi khi trích xuất dữ liệu có cấu trúc từ Excel: {str(e)}")
            return {}

    def find_data_by_keyword(self, file_path: str, keywords: List[str], case_sensitive: bool = False) -> Dict[
        str, List[Tuple[int, int, str]]]:
        """
        Tìm kiếm dữ liệu theo từ khóa trong file Excel

        Args:
            file_path: Đường dẫn đến file Excel
            keywords: Danh sách từ khóa cần tìm
            case_sensitive: Có phân biệt chữ hoa chữ thường không

        Returns:
            Dict[str, List[Tuple[int, int, str]]]: Từ điển kết quả với key là tên sheet và giá trị là danh sách (hàng, cột, giá trị)
        """
        results = {}
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_results = []

                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    for col_idx, cell_value in enumerate(row, start=1):
                        if cell_value is not None:
                            cell_str = str(cell_value)

                            for keyword in keywords:
                                if case_sensitive:
                                    if keyword in cell_str:
                                        sheet_results.append((row_idx, col_idx, cell_str))
                                else:
                                    if keyword.lower() in cell_str.lower():
                                        sheet_results.append((row_idx, col_idx, cell_str))

                if sheet_results:
                    results[sheet_name] = sheet_results

            workbook.close()

        except Exception as e:
            logger.error(f"Lỗi khi tìm kiếm từ khóa trong Excel: {str(e)}")

        return results

    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        Lấy danh sách tên các sheet trong file Excel

        Args:
            file_path: Đường dẫn đến file Excel

        Returns:
            List[str]: Danh sách tên các sheet
        """
        try:
            excel_data = pd.ExcelFile(file_path)
            return excel_data.sheet_names
        except Exception as e:
            logger.error(f"Lỗi khi lấy tên sheet: {str(e)}")
            return []

    def extract_tables(self, file_path: str) -> Dict[str, List[pd.DataFrame]]:
        """
        Trích xuất các bảng từ file Excel

        Args:
            file_path: Đường dẫn đến file Excel

        Returns:
            Dict[str, List[pd.DataFrame]]: Từ điển các bảng với key là tên sheet
        """
        tables = {}
        try:
            # Sử dụng openpyxl để xác định các vùng bảng
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Đọc toàn bộ sheet
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                # Tìm các khối dữ liệu liên tục không có ô trống
                # (Đây là thuật toán đơn giản để phát hiện bảng, có thể cần cải tiến cho các trường hợp phức tạp)
                mask = df.notna().all(axis=1)
                runs = self._identify_table_runs(mask)

                sheet_tables = []
                for start, end in runs:
                    if end - start > 1:  # Yêu cầu tối thiểu 2 hàng (tiêu đề + dữ liệu)
                        table_df = df.iloc[start:end + 1].reset_index(drop=True)
                        sheet_tables.append(table_df)

                if sheet_tables:
                    tables[sheet_name] = sheet_tables

            workbook.close()

        except Exception as e:
            logger.error(f"Lỗi khi trích xuất bảng từ Excel: {str(e)}")

        return tables

    def _identify_table_runs(self, mask: pd.Series) -> List[Tuple[int, int]]:
        """
        Xác định các đoạn liên tục của True trong mask

        Args:
            mask: Series boolean

        Returns:
            List[Tuple[int, int]]: Danh sách các cặp (start, end) của đoạn liên tục
        """
        runs = []
        start = None

        for i, value in enumerate(mask):
            if value and start is None:
                start = i
            elif not value and start is not None:
                runs.append((start, i - 1))
                start = None

        # Không quên đoạn cuối cùng
        if start is not None:
            runs.append((start, len(mask) - 1))

        return runs